import numpy as np 
import matplotlib.pyplot as plt 
from openpyxl import load_workbook

ws= load_workbook("X2.xlsx")["Sheet1"]
max_row= ws.max_row #miad mige chand ta satr dare
#print(max_row)
data = []
for i in range (1, max_row+1):
	data.append([ws[f"A{i}"].value,ws[f"B{i}"].value])

data = np.array(data) #tabdil mishe be matris 2 tayi az zoj moratab damane va bord
x=data[:,0] #kole adad sotoon aval ra midahad
y=data[:, 1]

dy=y[1:]-y[:-1]
dx=x[1:]-x[:-1]
ydot=y.copy()
ydot[1:]=dy/dx
plt.plot(x,ydot,label="2x")
plt.plot(x,y,label="x^2")
#plt.plot(x,y, label="sin")
plt.xlabel("x") #baraye mehvar x name x entekhab shode
plt.ylabel("y") #mesl khat 17 
plt.title("X^2 & 2x") #baraye kole nemoodar name entekhab shode, na mehvar ha

plt.show( )
plt.legend() #in khat agar baraye kole nemoodar esm entekhab mikonim bayaed bashad



#print(data[0])
	


